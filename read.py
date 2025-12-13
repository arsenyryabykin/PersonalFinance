
import sqlite3

import openpyxl

def get_data(cat_num: str, cat_name: str):
    # TODO: lalalala
    workbook = openpyxl.load_workbook("1.xlsx")
    sheet = workbook[str("Лист1")]

    content = []

    for i in range(2, 625):
        tmp = sheet.cell(row=i, column=1).value
        date = tmp.strftime('%Y-%m-%d')

        tmp_sum = sheet.cell(row=i, column=cat_num).value
        if tmp_sum is None:
            continue
        if isinstance(tmp_sum, str):
            b = tmp_sum[1:]
            bb = b.split('+')
            tmp_sum = sum([int(x) for x in bb])

        content.append((cat_name, tmp_sum, date))
    # print(content)
    return content

# a = get_data(5, 'products')
a = get_data(20, "save_acc_%")
print(a)